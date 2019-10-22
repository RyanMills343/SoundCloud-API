import pandas as pd
from pandas import ExcelWriter
from openpyxl import load_workbook
import numpy as np
import soundcloud
import datetime as dt


date = dt.datetime.today().strftime("%m-%d-%Y")
path = date+'.xlsx'
sheetname = 'Sheet1'

playlistnames = []
playlistids = []
playlistuserids = []
playlistuserfollowers = []
tracknames = []
trackids = []
trackplaycounts = []
trackfavcounts = []
trackuserids = []
trackfollowercounts = []

client = soundcloud.Client(client_id='1gZT1mMPZn4vbs2p7aDdnTTEC8r3wNWa')

def Main(user_url):
   playlistnames.clear()
   playlistids.clear()
   playlistuserids.clear()
   playlistuserfollowers.clear()
   tracknames.clear()
   trackids.clear()
   trackplaycounts.clear()
   trackfavcounts.clear()
   trackuserids.clear()
   trackfollowercounts.clear()
   
   #Fetch data metrics
   user = client.get('/resolve', url=user_url)
   playlists = client.get('/users/%d/playlists' % user.id, limit=5)
   for pl in playlists:
       playlistnames.append(pl.title)
       playlistids.append(pl.id)
       playlistuserids.append(pl.user_id)
       for track in pl.tracks:
           tracknames.append(track['title'])
           trackids.append(track['id'])
           try:
               trackfavcounts.append(track['favoritings_count'])
           except:
               continue   
           trackuserids.append(track['user_id'])
           
   for tracks in trackids:
      playback = client.get('/tracks/%d' % tracks)
      try:
           trackplaycounts.append(playback.playback_count)
      except:
           continue
   
   for users in playlistuserids:
       followers = client.get('/users/%d' % users)
       playlistuserfollowers.append(followers.followers_count)
   
   for users in trackuserids:
       followers = client.get('/users/%d' % users)
       trackfollowercounts.append(followers.followers_count)
       
def Excel(playlistnames, playlistids, playlistuserids, playlistuserfollowers, tracknames, trackids, trackplaycounts, trackfavcounts, trackuserids, trackfollowercounts):
   df = pd.DataFrame({'Playlist Name':pd.Series(playlistnames),
                      'Playlist ID':pd.Series(playlistids),
                      'Playlist UserID':pd.Series(playlistuserids),
                      'Playlist User Followers':pd.Series(playlistuserfollowers),
                      'Track Name':pd.Series(tracknames),
                      'Track ID':pd.Series(trackids),
                      'Track Plays':pd.Series(trackplaycounts),
                      'Track Favorites':pd.Series(trackfavcounts),
                      'Track UserID':pd.Series(trackuserids),
                      'UserID Followers':pd.Series(trackfollowercounts)})
    
   writer = ExcelWriter(path)
   df.to_excel(writer, sheetname ,index=False)
   writer.save()

def AddToExcel(playlistnames, playlistids, playlistuserids, playlistuserfollowers, tracknames, trackids, trackplaycounts, trackfavcounts, trackuserids, trackfollowercounts):
    wb = load_workbook(path)
    ws = wb[sheetname]
    row_count = ws.max_row
    df = pd.DataFrame({'Playlist Name':pd.Series(playlistnames),
                      'Playlist ID':pd.Series(playlistids),
                      'Playlist UserID':pd.Series(playlistuserids),
                      'Playlist User Followers':pd.Series(playlistuserfollowers),
                      'Track Name':pd.Series(tracknames),
                      'Track ID':pd.Series(trackids),
                      'Track Plays':pd.Series(trackplaycounts),
                      'Track Favorites':pd.Series(trackfavcounts),
                      'Track UserID':pd.Series(trackuserids),
                      'UserID Followers':pd.Series(trackfollowercounts)})
    
    writer = pd.ExcelWriter(path, engine='openpyxl')
    writer.book = load_workbook(path)
    writer.sheets = dict((ws.title,ws) for ws in writer.book.worksheets)
    df.to_excel(writer, sheet_name=sheetname, startrow=row_count ,index=False, header=False)
    writer.save()
    
def CleanExcel():
    df = pd.read_excel(path,
                       header=0,                      
                       na_values='null',              
                       index_col=None)
    df['Date'] = date
    cols = ['Playlist Name', 'Playlist ID', 'Playlist UserID', 'Playlist User Followers']
    df[cols] = df[cols].ffill()
    df['Track Plays'].replace('', np.nan, inplace=True)
    df.dropna(inplace=True)
    df.drop_duplicates(inplace=True)
    df.to_excel(excel_writer=path, sheet_name=sheetname, index=False)  

Main('https://soundcloud.com/runthetrap')
Excel(playlistnames, playlistids, playlistuserids, playlistuserfollowers, tracknames, trackids, trackplaycounts, trackfavcounts, trackuserids, trackfollowercounts)
Main('https://soundcloud.com/soundcloud-subs')
AddToExcel(playlistnames, playlistids, playlistuserids, playlistuserfollowers, tracknames, trackids, trackplaycounts, trackfavcounts, trackuserids, trackfollowercounts)
CleanExcel()
